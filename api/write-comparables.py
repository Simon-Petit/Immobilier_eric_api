"""
Vercel serverless function: POST /api/write-comparables
Accepts multipart/form-data: workbook (xlsx file), data (JSON string of filled configs).
Returns the modified workbook as xlsx download.
"""
from http.server import BaseHTTPRequestHandler
import json
import cgi
from io import BytesIO
import os
import sys
import traceback


sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl


def apply_comparables(workbook_bytes: bytes, data_str: str) -> bytes:
    if isinstance(data_str, bytes):
        data_str = data_str.decode("utf-8")

    payload = json.loads(data_str)
    print(f"payload type={type(payload).__name__}, len={len(payload) if isinstance(payload, list) else 'NOT A LIST'}")
    if isinstance(payload, list):
        for idx, comp in enumerate(payload):
            print(f"  comparable[{idx}]: type={type(comp).__name__}, len={len(comp) if isinstance(comp, list) else 'NOT A LIST'}, first_field={comp[0] if isinstance(comp, list) and comp else 'EMPTY'}")

    wb = openpyxl.load_workbook(BytesIO(workbook_bytes))
    print(f"workbook sheets: {wb.sheetnames}")

    for i, comparable in enumerate(payload):
        sheet_name = f"Comparable_{i + 1}"
        if sheet_name not in wb.sheetnames:
            print(f"  sheet '{sheet_name}' NOT FOUND in workbook, skipping")
            continue
        ws = wb[sheet_name]
        ws["C1"] = "Oui"
        written, skipped = 0, 0
        for field in comparable:
            if field.get("value") is not None:
                ws[field["cell"]] = field["value"]
                written += 1
            else:
                skipped += 1
        print(f"  sheet '{sheet_name}': written={written}, skipped_null={skipped}")

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    result = output.getvalue()
    print(f"output workbook size: {len(result)} bytes")
    return result


class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        if self.path != "/api/write-comparables" and not self.path.endswith("write-comparables"):
            self.send_error(404)
            return

        try:
            content_type = self.headers.get("Content-Type", "")
            content_length = int(self.headers.get("Content-Length", 0))
            print(f"content_type={content_type!r}, content_length={content_length}")

            if content_length == 0:
                self.send_error(400, "Missing body")
                return

            body = self.rfile.read(content_length)
            print(f"body bytes read: {len(body)}")

            env = {
                "REQUEST_METHOD": "POST",
                "CONTENT_TYPE": content_type,
                "CONTENT_LENGTH": str(content_length),
            }
            form = cgi.FieldStorage(fp=BytesIO(body), environ=env, keep_blank_values=True)

            try:
                form_keys = list(form.keys())
            except Exception as ke:
                form_keys = []
                print(f"form.keys() error: {ke}")
            print(f"form_keys={form_keys}")

            wb_field = form["workbook"] if "workbook" in form else None
            if "data" in form:
                data_field = form["data"]
                print("json_field=data")
            elif "comparables_array" in form:
                data_field = form["comparables_array"]
                print("json_field=comparables_array")
            else:
                data_field = None
                print(f"json_field=MISSING (got {form_keys})")

            print(f"wb_field is None: {wb_field is None}")
            print(f"data_field is None: {data_field is None}")
            if wb_field is not None:
                print(f"wb_field: type={type(wb_field).__name__}, has_file={hasattr(wb_field, 'file')}, filename={getattr(wb_field, 'filename', 'N/A')}")
            if data_field is not None:
                print(f"data_field: type={type(data_field).__name__}, has_value={hasattr(data_field, 'value')}")

            if wb_field is None:
                self.send_error(400, "Missing workbook field 'workbook'")
                return

            if data_field is None:
                self.send_error(400, "Missing JSON field 'data' or 'comparables_array'")
                return

            if not hasattr(wb_field, "file"):
                print(f"wb_field has no .file: {repr(wb_field)[:300]}")
                self.send_error(400, "workbook must be a file upload")
                return

            workbook_bytes = wb_field.file.read()
            print(f"workbook_bytes read: {len(workbook_bytes)} bytes")

            data_str = data_field.value if hasattr(data_field, "value") else data_field
            if isinstance(data_str, bytes):
                data_str = data_str.decode("utf-8")
            print(f"data_str length={len(data_str)}, preview={data_str[:300]!r}")

            result = apply_comparables(workbook_bytes, data_str)

            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="Evaluation_Immobiliere.xlsx"')
            self.send_header("Content-Length", str(len(result)))
            self.end_headers()
            self.wfile.write(result)

        except json.JSONDecodeError as e:
            print(f"JSONDecodeError: {repr(e)}")
            traceback.print_exc()
            self.send_error(400, f"Invalid JSON: {e}")
        except Exception as e:
            print(f"Exception: {repr(e)}")
            traceback.print_exc()
            self.send_error(500, str(e))
