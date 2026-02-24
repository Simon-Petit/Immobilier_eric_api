"""
quick_test.py — Local test harness for api/write-comparables.py

Usage (from vercel_deploy/ directory):
    python quick_test.py

Requirements:
    - A sample workbook file available locally (default: ../Evaluation_Immobiliere.xlsx)
    - Python dependencies from requirements.txt installed


This script does NOT run on Vercel; it's only for local debugging.
"""

import json
from io import BytesIO
from pathlib import Path

import openpyxl


def apply_comparables(workbook_bytes: bytes, data_str: str) -> bytes:
    """Same logic as api/write-comparables.py — inlined to avoid loading HTTP deps."""
    if isinstance(data_str, bytes):
        data_str = data_str.decode("utf-8")

    payload = json.loads(data_str)
    wb = openpyxl.load_workbook(BytesIO(workbook_bytes))

    for i, comparable in enumerate(payload):
        sheet_name = f"Comparable_{i + 1}"
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet {sheet_name} not found in workbook, skipping")
            continue
        ws = wb[sheet_name]
        ws["C1"] = "Oui"
        for field in comparable:
            if field.get("value") is not None:
                ws[field["cell"]] = field["value"]

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def main() -> None:
    root = Path(__file__).resolve().parent

    # 1) Where to read the workbook from
    workbook_path = (root / ".." / "Evaluation_Immobiliere.xlsx").resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found at: {workbook_path}")

    # 2) Define a minimal sample comparables_array matching the API contract
    #    (Update this structure as needed when debugging.)
    comparables_array = [
        [
            {"label": "Adresse (rue)", "cell": "C4", "value": "123 Rue Test"},
            {"label": "Ville", "cell": "C5", "value": "Ville-Test"},
            {"label": "No Centris", "cell": "C6", "value": "00000000"},
            {"label": "Statut", "cell": "C10", "value": "Vendu"},
        ]
    ]

    data_str = json.dumps(comparables_array, ensure_ascii=False)

    print(f"Loading workbook from: {workbook_path}")
    with workbook_path.open("rb") as f:
        workbook_bytes = f.read()

    print("Applying comparables locally (no HTTP/Vercel)...")
    result_bytes = apply_comparables(workbook_bytes, data_str)

    out_path = root / "Evaluation_Immobiliere_out.xlsx"
    with out_path.open("wb") as f:
        f.write(result_bytes)

    print(f"Done. Wrote modified workbook to: {out_path}")


if __name__ == "__main__":
    main()

